import glob
import os
from collections import Counter
from typing import Dict, List, Optional, Set, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from pandas.tseries.offsets import DateOffset


def read_source_excel(file_path: str, sheet_name: str) -> Optional[pd.DataFrame]:
    """
    Читает данные из исходного Excel файла.

    Args:
        file_path (str): Путь к Excel файлу
        sheet_name (str): Название листа для чтения

    Returns:
        Optional[pd.DataFrame]: DataFrame с данными или None в случае ошибки
    """
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return None


def map_support_level(warranty: str) -> str:
    """
    Преобразует значение гарантии в соответствующий уровень поддержки.

    Args:
        warranty (str): Исходное значение гарантии

    Returns:
        str: Уровень поддержки
    """
    if pd.isna(warranty):
        return "Не найдено"
    w = warranty.lower()
    if "notfound" in w:
        return "Не найдено"
    if "гарантия" in w:
        return "Гарантия"
    if w.startswith("base"):
        if "невозврат" in w:
            return "Базовый+невозврат"
        return "Базовый"
    if w.startswith("extended"):
        if "невозврат" in w:
            return "Расширенный+невозврат"
        return "Расширенный"
    if w.startswith("premium"):
        if "невозврат" in w:
            return "Премиум+невозврат"
        return "Премиум"
    return "Не найдено"


def format_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Форматирует данные из DataFrame под конечный формат.

    Args:
        df (pd.DataFrame): Исходный DataFrame с данными

    Returns:
        pd.DataFrame: Отформатированный DataFrame с нужными столбцами
    """
    # Срок: только цифра
    df["Срок"] = df["Warranty"].str.extract(r"(\d+)Y")[0]

    # SN OY: всегда строка, убираем .0 на конце
    def clean_sn(sn: Union[str, float, int]) -> str:
        """
        Очищает серийный номер от лишних символов.

        Args:
            sn (Union[str, float, int]): Исходный серийный номер

        Returns:
            str: Очищенный серийный номер
        """
        s = str(sn)
        if s.endswith(".0"):
            s = s[:-2]
        return s

    df["SN"] = df["SN"].apply(clean_sn)
    # Даты: формат YYYY-MM-DD, некорректные значения превращаем в пустую строку
    start_dates = pd.to_datetime(df["Начало гарантии"], errors="coerce")
    df["Дата начала"] = start_dates.dt.strftime("%Y-%m-%d").fillna("")

    # Дата окончания: дата начала + срок (в годах)
    def calc_end_date(row: pd.Series) -> str:
        """
        Рассчитывает дату окончания гарантии.

        Args:
            row (pd.Series): Строка данных с датой начала и сроком

        Returns:
            str: Дата окончания в формате YYYY-MM-DD или пустая строка в случае ошибки
        """
        try:
            start = pd.to_datetime(row["Дата начала"], errors="coerce")
            years = int(row["Срок"])
            if pd.isna(start) or pd.isna(years):
                return ""
            end = start + DateOffset(years=years)
            return end.strftime("%Y-%m-%d")
        except Exception:
            return ""

    df["Дата окончания"] = df.apply(calc_end_date, axis=1)
    # Форматируем уровень поддержки
    df["Уровень поддержки"] = df["Warranty"].apply(map_support_level)
    # Формируем итоговый датафрейм с нужными столбцами и переименованием
    result = pd.DataFrame(
        {
            "Наименование": df["Номенклатура"],
            "SN OY": df["SN"],
            "Уровень поддержки": df["Уровень поддержки"],
            "Срок": df["Срок"],
            "Дата начала": df["Дата начала"],
            "Дата окончания": df["Дата окончания"],
        }
    )
    return result


def check_duplicates(new_data: pd.DataFrame, target_file: str) -> pd.DataFrame:
    """
    Проверяет наличие дубликатов в целевом файле по SN OY.

    Args:
        new_data (pd.DataFrame): Новые данные для добавления
        target_file (str): Путь к целевому файлу

    Returns:
        pd.DataFrame: Объединенные данные без дубликатов
    """
    try:
        existing_data = pd.read_excel(target_file)
        combined_data = pd.concat([existing_data, new_data])
        # Оставляем только уникальные SN OY, оставляя первую встреченную строку
        combined_data = combined_data.drop_duplicates(subset=["SN OY"])
        return combined_data
    except FileNotFoundError:
        return new_data


def generate_analytics(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Генерирует аналитику по гарантиям.

    Args:
        df (pd.DataFrame): Основные данные о гарантиях

    Returns:
        Dict[str, pd.DataFrame]: Словарь с листами аналитики
    """
    print("Начинаем генерацию аналитики...")

    # Статистика по уровням поддержки
    print("Генерируем статистику по уровням поддержки...")
    support_stats = df["Уровень поддержки"].value_counts().reset_index()
    support_stats.columns = ["Уровень поддержки", "Количество"]
    support_stats["Процент"] = (support_stats["Количество"] / len(df) * 100).round(1)
    print(f"Статистика по уровням поддержки:\n{support_stats}")

    # Статистика по срокам
    print("Генерируем статистику по срокам...")
    term_stats = df["Срок"].value_counts().reset_index()
    term_stats.columns = ["Срок (лет)", "Количество"]
    term_stats["Процент"] = (term_stats["Количество"] / len(df) * 100).round(1)
    print(f"Статистика по срокам:\n{term_stats}")

    # Статистика по годам окончания гарантии
    print("Генерируем статистику по годам окончания...")
    df["Год окончания"] = pd.to_datetime(df["Дата окончания"]).dt.year
    year_stats = df["Год окончания"].value_counts().sort_index().reset_index()
    year_stats.columns = ["Год окончания", "Количество"]
    year_stats["Процент"] = (year_stats["Количество"] / len(df) * 100).round(1)
    print(f"Статистика по годам окончания:\n{year_stats}")

    # Общая статистика
    print("Генерируем общую статистику...")
    total_stats = pd.DataFrame(
        {
            "Показатель": [
                "Всего записей",
                "Уникальных SN",
                "Средний срок гарантии (лет)",
                "Минимальная дата начала",
                "Максимальная дата окончания",
            ],
            "Значение": [
                len(df),
                df["SN OY"].nunique(),
                df["Срок"].astype(float).mean().round(1),
                df["Дата начала"].min(),
                df["Дата окончания"].max(),
            ],
        }
    )
    print(f"Общая статистика:\n{total_stats}")

    result = {
        "Общая статистика": total_stats,
        "По уровням поддержки": support_stats,
        "По срокам": term_stats,
        "По годам окончания": year_stats,
    }

    print("Аналитика успешно сгенерирована")
    return result


def save_to_excel(df: pd.DataFrame, target_file: str) -> None:
    """
    Сохраняет данные в целевой Excel файл.

    Args:
        df (pd.DataFrame): Данные для сохранения
        target_file (str): Путь к целевому файлу
    """
    try:
        print("Начинаем сохранение данных в Excel...")
        # Создаем Excel writer
        with pd.ExcelWriter(target_file, engine="openpyxl") as writer:
            print("Сохраняем основные данные...")
            # Сохраняем основные данные
            df.to_excel(writer, sheet_name="Гарантии", index=False)

            print("Генерируем аналитику...")
            # Генерируем и сохраняем аналитику
            analytics = generate_analytics(df)
            print(f"Создаем листы аналитики: {list(analytics.keys())}")

            for sheet_name, data in analytics.items():
                print(f"Сохраняем лист '{sheet_name}'...")
                data.to_excel(writer, sheet_name=sheet_name, index=False)

                # Получаем лист для форматирования
                worksheet = writer.sheets[sheet_name]

                # Форматируем заголовки
                for cell in worksheet[1]:
                    cell.font = cell.font.copy(bold=True)
                    cell.fill = PatternFill(
                        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                    )

                # Автоматическая ширина столбцов
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[column[0].column_letter].width = (
                        adjusted_width
                    )

                print(f"Лист '{sheet_name}' отформатирован")

        print(f"Данные успешно сохранены в файл {target_file}")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
        raise  # Добавляем raise для отладки


def colorize_excel(target_file: str) -> None:
    """
    Применяет цветовое форматирование к Excel файлу.

    Args:
        target_file (str): Путь к целевому файлу
    """
    # Цвета для уровня поддержки
    support_colors: Dict[str, str] = {
        "Базовый": "ADD8E6",  # светло-синий
        "Базовый+невозврат": "87CEEB",  # синий
        "Гарантия": "90EE90",  # зелёный
        "Премиум": "FFFF99",  # жёлтый
        "Премиум+невозврат": "FFA500",  # оранжевый
        "Расширенный": "DDA0DD",  # фиолетовый
        "Расширенный+невозврат": "FF6347",  # красный
        "Не найдено": "D3D3D3",  # серый
    }
    # Цвета для срока
    term_colors: Dict[str, str] = {
        "1": "CCFFCC",  # светло-зелёный
        "3": "FFFFCC",  # светло-жёлтый
        "5": "FFE4B5",  # светло-оранжевый
    }
    # Цвет для дубликатов SN OY
    duplicate_sn_color: str = "FFC7CE"  # светло-красный

    wb = load_workbook(target_file)
    ws: Worksheet = wb["Гарантии"]  # Теперь явно указываем лист "Гарантии"

    # Найдём индексы нужных столбцов по заголовкам
    header: Dict[str, int] = {
        cell.value: idx
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1)
    }
    col_support = header.get("Уровень поддержки")
    col_term = header.get("Срок")
    col_sn = header.get("SN OY")

    # Соберём все значения SN OY для поиска дубликатов
    sn_values: List[str] = [
        row[col_sn - 1].value for row in ws.iter_rows(min_row=2) if col_sn
    ]
    sn_counts = Counter(sn_values)
    duplicate_sns: Set[str] = {sn for sn, count in sn_counts.items() if count > 1}

    # Применяем цвета
    for row in ws.iter_rows(min_row=2):
        # Уровень поддержки
        if col_support:
            val = row[col_support - 1].value
            color = support_colors.get(val, None)
            if color:
                row[col_support - 1].fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
        # Срок
        if col_term:
            val = str(row[col_term - 1].value)
            color = term_colors.get(val, None)
            if color:
                row[col_term - 1].fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
        # Дубликаты SN OY
        if col_sn:
            val = row[col_sn - 1].value
            if val in duplicate_sns:
                row[col_sn - 1].fill = PatternFill(
                    start_color=duplicate_sn_color,
                    end_color=duplicate_sn_color,
                    fill_type="solid",
                )
    wb.save(target_file)


def main() -> None:
    """
    Основная функция программы.
    Обрабатывает Excel файлы в папке input и создает отформатированный файл в output.
    """
    # Пути к файлам
    # Скрипт обрабатывает все Excel-файлы в папке input
    input_dir: str = "input"
    output_dir: str = "output"
    sheet_name: str = "Гарантия"
    target_file: str = f"{output_dir}/target.xlsx"

    # Создаем папки input и output если они не существуют
    for directory in [input_dir, output_dir]:
        if not os.path.exists(directory):
            try:
                os.makedirs(directory)
                print(f"Создана папка {directory}")
            except Exception as e:
                print(f"Ошибка при создании папки {directory}: {e}")
                return

    # Удаляем существующий target.xlsx если он есть
    if os.path.exists(target_file):
        try:
            os.remove(target_file)
            print(f"Удален существующий файл {target_file}")
        except Exception as e:
            print(f"Ошибка при удалении файла {target_file}: {e}")
            return

    # Собираем все Excel-файлы в input
    input_files: List[str] = glob.glob(f"{input_dir}/*.xls*")
    if not input_files:
        print(f"Нет Excel-файлов в папке {input_dir}!")
        return

    # Создаём новый файл с нужной шапкой
    columns: List[str] = [
        "Наименование",
        "SN OY",
        "Уровень поддержки",
        "Срок",
        "Дата начала",
        "Дата окончания",
    ]
    pd.DataFrame(columns=columns).to_excel(target_file, index=False)
    print(f"Создан новый файл {target_file} с нужной структурой.")

    # Читаем и объединяем данные из всех файлов
    all_data: List[pd.DataFrame] = []
    for file in input_files:
        print(f"Чтение файла: {file}")
        source_data = read_source_excel(file, sheet_name)
        if source_data is None:
            print(f"Пропущен файл {file} (нет листа '{sheet_name}' или ошибка чтения)")
            continue
        formatted_data = format_data(source_data)
        all_data.append(formatted_data)

    if not all_data:
        print("Нет данных для обработки!")
        return

    # Объединяем все данные в один DataFrame
    combined_data: pd.DataFrame = pd.concat(all_data, ignore_index=True)

    # Проверяем дубликаты и получаем финальный набор данных
    final_data: pd.DataFrame = check_duplicates(combined_data, target_file)

    # Сохраняем результат
    save_to_excel(final_data, target_file)

    # Цветовое форматирование
    colorize_excel(target_file)


if __name__ == "__main__":
    main()
